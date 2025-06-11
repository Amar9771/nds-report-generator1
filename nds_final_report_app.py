import streamlit as st
import pandas as pd
import io
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook

# 🎯 Page Setup
st.set_page_config(page_title="📄 NDS Final Report Generator", layout="centered")

# 🔰 Optional Logo
# st.image("https://your-logo-url.com/logo.png", width=120)

# 🏷️ Title & Description
st.title("📄 NDS Final Report Generator")
st.markdown("This tool generates a **3-sheet formatted Excel report** based on monthly NDS submissions.")

# 📘 Sidebar Help
with st.sidebar:
    st.header("ℹ️ Instructions")
    st.markdown("""
    1. Upload all 4 required Excel files:
        - Master File
        - March File
        - April File
        - May File  
    2. Click **'Generate Report'**
    3. Download the final Excel report  
    """)

    # 📥 Sample File Download (optional: make sure 'sample_format.xlsx' exists)
    try:
        with open("sample_format.xlsx", "rb") as f:
            st.download_button("📥 Download Sample Format", f, file_name="sample_format.xlsx")
    except FileNotFoundError:
        st.info("Sample format file not found.")

# 📁 File Uploads
uploaded_master = st.file_uploader("🔹 Upload All Entity Master File", type="xlsx")
uploaded_march = st.file_uploader("🔸 Upload March File", type="xlsx")
uploaded_april = st.file_uploader("🔸 Upload April File", type="xlsx")
uploaded_may = st.file_uploader("🔸 Upload May File", type="xlsx")

def get_ids(df):
    return set(df['Organization']) if df is not None else set()

def apply_formatting(ws):
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    for col in ws.iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.font = Font(bold=True)
            cell.fill = header_fill
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

if all([uploaded_master, uploaded_march, uploaded_april, uploaded_may]):
    if st.button("📊 Generate Report"):

        # Read files
        df_master = pd.read_excel(uploaded_master)
        df_march = pd.read_excel(uploaded_march)
        df_april = pd.read_excel(uploaded_april)
        df_may = pd.read_excel(uploaded_may)

        # Get IDs
        march_ids = get_ids(df_march)
        april_ids = get_ids(df_april)
        may_ids = get_ids(df_may)

        # Sheet 1: Entity List
        entity_df = df_master[["Organization", "Name"]].copy()

        # Sheet 2: Last 3 Months Not Submitted
        last3_df = df_master[
            df_master["Organization"].apply(lambda x: x not in march_ids and x not in april_ids and x not in may_ids)
        ][["Organization", "Name"]].copy()

        # Sheet 3: Summary
        def missing_count(org_id):
            return sum([org_id not in march_ids, org_id not in april_ids, org_id not in may_ids])

        summary_df = df_master[["Organization", "Name"]].copy()
        summary_df["Months"] = summary_df["Organization"].apply(missing_count)
        summary_df = summary_df[summary_df["Months"] > 0]

        # 📒 Excel Workbook
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Entity_SelfList"
        for r in dataframe_to_rows(entity_df, index=False, header=True):
            ws1.append(r)
        apply_formatting(ws1)

        ws2 = wb.create_sheet("Last 3 Months Not Submitted")
        for r in dataframe_to_rows(last3_df, index=False, header=True):
            ws2.append(r)
        apply_formatting(ws2)

        ws3 = wb.create_sheet("Summary")
        for r in dataframe_to_rows(summary_df, index=False, header=True):
            ws3.append(r)
        apply_formatting(ws3)

        # 💾 Save and Download
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        st.success("✅ Report generated successfully with formatting!")
        st.download_button(
            label="📥 Download Formatted Excel Report",
            data=output,
            file_name="NDS_Final_Formatted_Report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        st.subheader("📋 Preview: 'Last 3 Months Not Submitted'")
        st.dataframe(last3_df.head(), use_container_width=True)

else:
    st.info("📂 Please upload all 4 required Excel files to continue.")
